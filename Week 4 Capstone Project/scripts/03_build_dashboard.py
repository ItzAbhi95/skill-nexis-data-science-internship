"""
Week 4 Capstone - End-to-End Analysis Case Study
Step 3: Excel Dashboard (Power BI-style, formula-driven)

Builds a formatted Excel workbook with:
 - Raw data sheet
 - Summary tables (driven by SUMIFS formulas -> recalculate if data changes)
 - KPI cards
 - Bar / Line / Pie charts for business-metric storytelling
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList

df = pd.read_csv("../outputs/cleaned_sales_data.csv")

wb = Workbook()

# -----------------------------------------------------------------
# Styling constants
# -----------------------------------------------------------------
NAVY = "1F3864"
GOLD = "C99A2E"
WHITE = "FFFFFF"
LIGHT = "F2F2F2"
FONT_NAME = "Arial"

header_font = Font(name=FONT_NAME, bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT_NAME, bold=True, color=NAVY, size=16)
kpi_label_font = Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
kpi_value_font = Font(name=FONT_NAME, bold=True, color=WHITE, size=18)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===================================================================
# SHEET 1: Raw Data
# ===================================================================
ws_data = wb.active
ws_data.title = "Data"

for c_idx, col in enumerate(df.columns, start=1):
    cell = ws_data.cell(row=1, column=c_idx, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for r_idx, row in enumerate(df.itertuples(index=False), start=2):
    for c_idx, val in enumerate(row, start=1):
        ws_data.cell(row=r_idx, column=c_idx, value=val)

for c_idx, col in enumerate(df.columns, start=1):
    ws_data.column_dimensions[get_column_letter(c_idx)].width = max(12, len(col) + 2)

n_rows = len(df) + 1  # includes header
col_letter = {col: get_column_letter(i + 1) for i, col in enumerate(df.columns)}

ws_data.freeze_panes = "A2"

# ===================================================================
# SHEET 2: Summary tables (formula-driven, feed the charts)
# ===================================================================
ws_sum = wb.create_sheet("Summary")

def style_table_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

# --- Table 1: Sales & Profit by Segment ---
ws_sum["A1"] = "Sales & Profit by Segment"
ws_sum["A1"].font = Font(name=FONT_NAME, bold=True, size=13, color=NAVY)

segments = sorted(df["Segment"].unique())
ws_sum.append([])
headers = ["Segment", "Total Sales", "Total Profit", "Profit Margin %"]
ws_sum.append(headers)
style_table_header(ws_sum, 3, 1, 4)

seg_start_row = 4
for i, seg in enumerate(segments):
    r = seg_start_row + i
    ws_sum.cell(row=r, column=1, value=seg).border = border
    ws_sum.cell(row=r, column=2,
                value=f"=SUMIF(Data!${col_letter['Segment']}$2:${col_letter['Segment']}${n_rows},A{r},"
                      f"Data!${col_letter['Sales']}$2:${col_letter['Sales']}${n_rows})").border = border
    ws_sum.cell(row=r, column=3,
                value=f"=SUMIF(Data!${col_letter['Segment']}$2:${col_letter['Segment']}${n_rows},A{r},"
                      f"Data!${col_letter['Profit']}$2:${col_letter['Profit']}${n_rows})").border = border
    ws_sum.cell(row=r, column=4, value=f"=C{r}/B{r}").border = border
    ws_sum.cell(row=r, column=2).number_format = "$#,##0"
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
    ws_sum.cell(row=r, column=4).number_format = "0.0%"
seg_end_row = seg_start_row + len(segments) - 1

# --- Table 2: Sales & Profit by Country ---
country_title_row = seg_end_row + 3
ws_sum.cell(row=country_title_row, column=1, value="Sales & Profit by Country").font = Font(
    name=FONT_NAME, bold=True, size=13, color=NAVY)
header_row2 = country_title_row + 2
ws_sum.cell(row=header_row2, column=1, value="Country")
ws_sum.cell(row=header_row2, column=2, value="Total Sales")
ws_sum.cell(row=header_row2, column=3, value="Total Profit")
style_table_header(ws_sum, header_row2, 1, 3)

countries = sorted(df["Country"].unique())
country_start_row = header_row2 + 1
for i, ctry in enumerate(countries):
    r = country_start_row + i
    ws_sum.cell(row=r, column=1, value=ctry).border = border
    ws_sum.cell(row=r, column=2,
                value=f"=SUMIF(Data!${col_letter['Country']}$2:${col_letter['Country']}${n_rows},A{r},"
                      f"Data!${col_letter['Sales']}$2:${col_letter['Sales']}${n_rows})").border = border
    ws_sum.cell(row=r, column=3,
                value=f"=SUMIF(Data!${col_letter['Country']}$2:${col_letter['Country']}${n_rows},A{r},"
                      f"Data!${col_letter['Profit']}$2:${col_letter['Profit']}${n_rows})").border = border
    ws_sum.cell(row=r, column=2).number_format = "$#,##0"
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
country_end_row = country_start_row + len(countries) - 1

# --- Table 3: Sales & Profit by Product ---
product_title_row = country_end_row + 3
ws_sum.cell(row=product_title_row, column=1, value="Sales & Profit by Product").font = Font(
    name=FONT_NAME, bold=True, size=13, color=NAVY)
header_row3 = product_title_row + 2
ws_sum.cell(row=header_row3, column=1, value="Product")
ws_sum.cell(row=header_row3, column=2, value="Total Sales")
ws_sum.cell(row=header_row3, column=3, value="Total Profit")
style_table_header(ws_sum, header_row3, 1, 3)

products = sorted(df["Product"].unique())
product_start_row = header_row3 + 1
for i, prod in enumerate(products):
    r = product_start_row + i
    ws_sum.cell(row=r, column=1, value=prod).border = border
    ws_sum.cell(row=r, column=2,
                value=f"=SUMIF(Data!${col_letter['Product']}$2:${col_letter['Product']}${n_rows},A{r},"
                      f"Data!${col_letter['Sales']}$2:${col_letter['Sales']}${n_rows})").border = border
    ws_sum.cell(row=r, column=3,
                value=f"=SUMIF(Data!${col_letter['Product']}$2:${col_letter['Product']}${n_rows},A{r},"
                      f"Data!${col_letter['Profit']}$2:${col_letter['Profit']}${n_rows})").border = border
    ws_sum.cell(row=r, column=2).number_format = "$#,##0"
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
product_end_row = product_start_row + len(products) - 1

# --- Table 4: Monthly trend (Year-Month) ---
trend_title_row = product_end_row + 3
ws_sum.cell(row=trend_title_row, column=1, value="Monthly Sales & Profit Trend").font = Font(
    name=FONT_NAME, bold=True, size=13, color=NAVY)
header_row4 = trend_title_row + 2
ws_sum.cell(row=header_row4, column=1, value="Period")
ws_sum.cell(row=header_row4, column=2, value="Total Sales")
ws_sum.cell(row=header_row4, column=3, value="Total Profit")
style_table_header(ws_sum, header_row4, 1, 3)

monthly = (df.groupby(["Year", "Month Number", "Month Name"])
             .size().reset_index()[["Year", "Month Number", "Month Name"]]
             .sort_values(["Year", "Month Number"]))
monthly["Period"] = monthly["Month Name"].str[:3] + "-" + monthly["Year"].astype(str)

trend_start_row = header_row4 + 1
for i, row in enumerate(monthly.itertuples(index=False)):
    r = trend_start_row + i
    ws_sum.cell(row=r, column=1, value=row.Period).border = border
    # SUMIFS on Year AND Month Number
    ws_sum.cell(row=r, column=2,
                value=(f"=SUMIFS(Data!${col_letter['Sales']}$2:${col_letter['Sales']}${n_rows},"
                       f"Data!${col_letter['Year']}$2:${col_letter['Year']}${n_rows},{row.Year},"
                       f"Data!${col_letter['Month Number']}$2:${col_letter['Month Number']}${n_rows},{row._1})")
                ).border = border
    ws_sum.cell(row=r, column=3,
                value=(f"=SUMIFS(Data!${col_letter['Profit']}$2:${col_letter['Profit']}${n_rows},"
                       f"Data!${col_letter['Year']}$2:${col_letter['Year']}${n_rows},{row.Year},"
                       f"Data!${col_letter['Month Number']}$2:${col_letter['Month Number']}${n_rows},{row._1})")
                ).border = border
    ws_sum.cell(row=r, column=2).number_format = "$#,##0"
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
trend_end_row = trend_start_row + len(monthly) - 1

for c in range(1, 5):
    ws_sum.column_dimensions[get_column_letter(c)].width = 20

# ===================================================================
# SHEET 3: Dashboard (KPI cards + charts)
# ===================================================================
ws_dash = wb.create_sheet("Dashboard", 0)  # make it the first visible sheet
ws_dash.sheet_view.showGridLines = False

ws_dash["B2"] = "SALES PERFORMANCE DASHBOARD"
ws_dash["B2"].font = title_font
ws_dash["B3"] = "Business Metrics Overview — Sample Sales Dataset (2013–2014)"
ws_dash["B3"].font = Font(name=FONT_NAME, italic=True, color="666666", size=10)

# --- KPI cards (row 5-8) ---
kpi_defs = [
    ("Total Sales", f"=SUM(Data!{col_letter['Sales']}2:{col_letter['Sales']}{n_rows})", "$#,##0"),
    ("Total Profit", f"=SUM(Data!{col_letter['Profit']}2:{col_letter['Profit']}{n_rows})", "$#,##0"),
    ("Overall Profit Margin",
     f"=SUM(Data!{col_letter['Profit']}2:{col_letter['Profit']}{n_rows})/"
     f"SUM(Data!{col_letter['Sales']}2:{col_letter['Sales']}{n_rows})", "0.0%"),
    ("Total Orders", f"=COUNTA(Data!A2:A{n_rows})", "#,##0"),
]

kpi_col_start = 2
kpi_width = 3
for i, (label, formula, numfmt) in enumerate(kpi_defs):
    c0 = kpi_col_start + i * kpi_width
    c1 = c0 + kpi_width - 1
    ws_dash.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c1)
    ws_dash.merge_cells(start_row=6, start_column=c0, end_row=7, end_column=c1)
    label_cell = ws_dash.cell(row=5, column=c0, value=label)
    label_cell.font = kpi_label_font
    label_cell.alignment = Alignment(horizontal="center")
    val_cell = ws_dash.cell(row=6, column=c0, value=formula)
    val_cell.font = kpi_value_font
    val_cell.number_format = numfmt
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    fill_color = NAVY if i % 2 == 0 else GOLD
    for rr in (5, 6, 7):
        for cc in range(c0, c1 + 1):
            ws_dash.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=fill_color)

for c in range(2, 15):
    ws_dash.column_dimensions[get_column_letter(c)].width = 9
for r in (5, 6, 7):
    ws_dash.row_dimensions[r].height = 22

# --- Chart 1: Profit by Segment (Bar) ---
bar1 = BarChart()
bar1.title = "Profit by Segment"
bar1.y_axis.title = "Profit ($)"
bar1.style = 10
data_ref = Reference(ws_sum, min_col=3, min_row=3, max_row=seg_end_row)
cats_ref = Reference(ws_sum, min_col=1, min_row=seg_start_row, max_row=seg_end_row)
bar1.add_data(data_ref, titles_from_data=True)
bar1.set_categories(cats_ref)
bar1.width, bar1.height = 15, 9
ws_dash.add_chart(bar1, "B10")

# --- Chart 2: Sales by Country (Bar) ---
bar2 = BarChart()
bar2.title = "Sales by Country"
bar2.y_axis.title = "Sales ($)"
bar2.style = 11
data_ref2 = Reference(ws_sum, min_col=2, min_row=header_row2, max_row=country_end_row)
cats_ref2 = Reference(ws_sum, min_col=1, min_row=country_start_row, max_row=country_end_row)
bar2.add_data(data_ref2, titles_from_data=True)
bar2.set_categories(cats_ref2)
bar2.width, bar2.height = 15, 9
ws_dash.add_chart(bar2, "H10")

# --- Chart 3: Monthly Sales Trend (Line) ---
line1 = LineChart()
line1.title = "Monthly Sales Trend"
line1.y_axis.title = "Sales ($)"
line1.style = 12
data_ref3 = Reference(ws_sum, min_col=2, min_row=header_row4, max_row=trend_end_row)
cats_ref3 = Reference(ws_sum, min_col=1, min_row=trend_start_row, max_row=trend_end_row)
line1.add_data(data_ref3, titles_from_data=True)
line1.set_categories(cats_ref3)
line1.width, line1.height = 22, 9
ws_dash.add_chart(line1, "B28")

# --- Chart 4: Sales Mix by Product (Pie) ---
pie1 = PieChart()
pie1.title = "Sales Mix by Product"
data_ref4 = Reference(ws_sum, min_col=2, min_row=header_row3, max_row=product_end_row)
cats_ref4 = Reference(ws_sum, min_col=1, min_row=product_start_row, max_row=product_end_row)
pie1.add_data(data_ref4, titles_from_data=True)
pie1.set_categories(cats_ref4)
pie1.dataLabels = DataLabelList()
pie1.dataLabels.showPercent = True
pie1.width, pie1.height = 15, 9
ws_dash.add_chart(pie1, "O28")

wb.active = 0  # open on Dashboard sheet
ws_dash.print_area = "A1:X50"
ws_dash.page_setup.orientation = "landscape"
ws_dash.page_setup.fitToWidth = 1
ws_dash.page_setup.fitToHeight = 0
ws_dash.sheet_properties.pageSetUpPr.fitToPage = True

OUT_PATH = "../outputs/Sales_Performance_Dashboard.xlsx"
wb.save(OUT_PATH)
print(f"Dashboard workbook saved -> {OUT_PATH}")
